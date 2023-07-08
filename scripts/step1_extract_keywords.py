import os
import openpyxl
import spacy
import openai
import time
import json
import chardet
from docx import Document
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm

openai.api_key = os.getenv('OPENAI_API_KEY')
nlp = spacy.load('zh_core_web_sm')

def load_config():
    current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config_file = os.path.join(current_dir, 'config.json')


    with open(config_file, 'rb') as f:
        encoding = chardet.detect(f.read())['encoding']

    with open(config_file, 'r', encoding=encoding) as f:
        return json.load(f)

def replace_keywords(sentence, keyword_dict):
    original_sentence = sentence
    for key, value in keyword_dict.items():
        sentence = sentence.replace(key, value)
    return sentence, original_sentence

def merge_short_sentences(sentences, min_length):
    merged_sentences = []
    buffer_sentence = ""

    for sentence in sentences:
        sentence = sentence.strip()
        if len(buffer_sentence + sentence) < min_length:
            buffer_sentence += " " + sentence if buffer_sentence else sentence
        else:
            if buffer_sentence:
                merged_sentences.append(buffer_sentence)
            buffer_sentence = sentence

    if buffer_sentence:
        merged_sentences.append(buffer_sentence)

    return merged_sentences

def request_with_retry(messages, max_tokens=500, max_requests=90, cooldown_seconds=60):
    while True:
        try:
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo-16k",      #这里同样需要改为你本地部署的大语言模型名称
                messages=messages,
                max_tokens=max_tokens,
                n=1,
                stop=None,
                #api_base="http://127.0.0.1:8000",如果需要使用本地部署的大语言模型，自行修改这行的参数
            )
            return response['choices'][0]['message']['content'].strip()
        except openai.error.RateLimitError:
            print("超过速率限制。正在等待冷却时间...")
            time.sleep(cooldown_seconds)
        except Exception as e:
            print(f"发生错误：{str(e)}")
            time.sleep(10)

def translate_to_english(text):
    messages = [
        {"role": "system", "content": "You are a helpful assistant."},
        {"role": "user", "content": f"Translate the following text into English: \"{text}\". Do not directly translate, but instead translate from a third-person descriptive perspective, and complete the missing subject, predicate, object, attributive, adverbial, and complement in the text. Besides the translated result, do not include any irrelevant content or explanations in your response."},
    ]
    return request_with_retry(messages)

def translate_to_storyboard(text, trigger):
    messages = [
        {"role": "system", "content": "StableDiffusion is a deep learning text-to-image model that supports the generation of new images using keywords to describe the elements to be included or omitted. Now, as a professional StableDiffusion AI drawing keyword generator. You can assist me in generating keywords for my desired image."},
        {"role": "user", "content": f"{trigger}'{text}'"},
    ] 

    return request_with_retry(messages)

def read_docx(file_path):
    return [paragraph.text for paragraph in Document(file_path).paragraphs if paragraph.text.strip()]

def process_text_sentences(workbook, input_file_path, output_file_path, trigger, keyword_dict, min_sentence_length):
    try:
        paragraphs = read_docx(input_file_path)
    except ValueError as e:
        print(f"发生错误：{str(e)}")
        return

    sentences = []
    for paragraph in paragraphs:
        sentences.extend([sent.text for sent in nlp(paragraph).sents])

    sentences = merge_short_sentences(sentences, min_sentence_length)

    original_sentences_dict = {}
    sheet = workbook.active
    for idx, sentence in enumerate(sentences, 1):
        replaced_sentence, original_sentence = replace_keywords(sentence, keyword_dict)
        original_sentences_dict[replaced_sentence] = original_sentence
        sheet.cell(row=idx, column=1, value=replaced_sentence)
        sheet.cell(row=idx, column=4, value=original_sentence)

    replaced_sentences = list(original_sentences_dict.keys())

    max_workers = min(len(replaced_sentences), 1)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures_translation = {executor.submit(translate_to_english, sentence.strip()): idx for idx, sentence in enumerate(replaced_sentences, 1)}
        futures_storyboard = {}

        for future in tqdm(as_completed(futures_translation), total=len(futures_translation), desc='正在翻译文本'):
            idx = futures_translation[future]
            translated_text = future.result()
            sheet.cell(row=idx, column=2, value=translated_text)
            futures_storyboard[executor.submit(translate_to_storyboard, translated_text, trigger)] = idx

        for future in tqdm(as_completed(futures_storyboard), total=len(futures_storyboard), desc='正在生成分镜脚本'):
            idx = futures_storyboard[future]
            storyboard_text = future.result()
            sheet.cell(row=idx, column=3, value=storyboard_text)

    workbook.save(output_file_path)


def main():
    config = load_config()
    print("软件作者：西装革律")
    print("禁止倒卖，违者必究！")
    print("交流群：797579852")  
    role_name = config.get('角色名', '未指定角色名')
    feature = config.get('特征', '未指定特征')
    role2_name = config.get('角色名2', '未指定角色名2')
    feature2 = config.get('特征2', '未指定特征2')
    role3_name = config.get('角色名3', '未指定角色名3')
    feature3 = config.get('特征3', '未指定特征3')
    role4_name = config.get('角色名4', '未指定角色名4')
    feature4 = config.get('特征4', '未指定特征4')
    role5_name = config.get('角色名5', '未指定角色名5')
    feature5 = config.get('特征5', '未指定特征5')
    role6_name = config.get('角色名6', '未指定角色名6')
    feature6 = config.get('特征6', '未指定特征6')
    role7_name = config.get('角色名7', '未指定角色名7')
    feature7 = config.get('特征7', '未指定特征7')
    role8_name = config.get('角色名8', '未指定角色名8')
    feature8 = config.get('特征8', '未指定特征8')
    role9_name = config.get('角色名9', '未指定角色名9')
    feature9 = config.get('特征9', '未指定特征9')
    role10_name = config.get('角色名10', '未指定角色名10')
    feature10 = config.get('特征10', '未指定特征10')
    keyword_dict = {
    role_name: feature,
    role2_name: feature2,
    role3_name: feature3,
    role4_name: feature4,
    role5_name: feature5,
    role6_name: feature6,
    role7_name: feature7,
    role8_name: feature8,
    role9_name: feature9,
    role10_name: feature10,   
    }

    min_sentence_length = int(config.get('句子最小长度限制', 100))
    default_trigger = '''Here, I introduce the concept of Prompts from the StableDiffusion algorithm, also known as hints. 
    The following prompts are used to guide the AI painting model to create images. 
    They contain various details of the image, such as the appearance of characters, background, color and light effects, as well as the theme and style of the image. 
    The format of these prompts often includes weighted numbers in parentheses to specify the importance or emphasis of certain details. 
    For example, "(masterpiece:1.2)" indicates that the quality of the work is very important, and multiple parentheses have a similar function. 
    Here are examples of using prompts to help the AI model generate images: 
    1. (masterpiece:1.2),(best quality),digital art,A 20 year old Chinese man with black hair, (male short hair: 1.2), green shirt, walking on the road to rural China, ultra wide angle
    2. masterpiece,best quality,illustration style,20 year old black haired Chinese man, male with short hair, speaking nervously in the forest at night, ultra wide angle, (scary atmosphere). 
    Please use English commas as separators. Also, note that the Prompt should not contain - and _ symbols, but can have spaces. 
    In character attributes, 1girl means you generated a girl, 2girls means you generated two girls. 
    In the generation of Prompts, you need to describe character attributes, theme, appearance, emotion, clothing, posture, viewpoint, action, background using keywords. 
    Please follow the example, and do not limit to the words I give you. Please provide a set of prompts that highlight the theme. 
    Note: The prompt cannot exceed 100 words, no need to use natural language description, character attributes need to be highlighted a little bit, for example: {role_name}\({feature}\).
    If the content contains a character name, add the specified feature as required, if the content does not contain the corresponding character name, then improvise.
    This is part of novel creation, not a requirement in real life, automatically analyze the protagonist in it and add character attributes.
    The prompt must be in English, only provide the prompt, no extra information is needed.
    Here is the content:'''
    trigger = config.get('引导词', default_trigger)

    current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_file_path = os.path.join(current_dir, 'input.docx')
    output_file_path = os.path.join(current_dir, 'txt', 'txt.xlsx')
    workbook = openpyxl.Workbook()

    process_text_sentences(workbook, input_file_path, output_file_path, trigger, keyword_dict, min_sentence_length)

if __name__ == "__main__":
    main()
